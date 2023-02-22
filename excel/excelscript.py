import requests
from PIL import Image
from io import BytesIO
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('example.xlsx')

# Select the active worksheet
ws = wb.active
i=0
# Loop through each row in the first column
for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
    # Get the URL from the first column
    url = row[0]
    i+=1

    # Send a request to retrieve the image
    response = requests.get(url)
    
    # Open the image using Pillow
    img = Image.open(BytesIO(response.content))
    
    # Get the resolution of the image
    resolution = str(img.size)
    
    # Write the resolution to the second column
    ws.cell(row=i, column=2, value=resolution)

# Save the modified Excel file
wb.save('example.xlsx')
